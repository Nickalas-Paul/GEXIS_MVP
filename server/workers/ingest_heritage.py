"""
Ingest Heritage Foundation Index of Economic Freedom into raw_indicators.

Downloads yearly Excel workbooks from the Heritage download hub
(https://indexdotnet.azurewebsites.net/index/download) and caches them under
server/workers/data/.

Historical coverage note (Step 5B):
  - The published Excel URL pattern
    ``…/index/excel/{year}/index{year}_data.xlsx`` only resolves for 2022 and
    2023 (older years return 404).
  - Each workbook is single-year for the component scores we need (plus a prior
    overall score column on 2023 — not used for components).
  - A full multi-year bulk download exists via the interactive “all country
    scores” UI, but is not a straightforward machine-readable endpoint; within
    the investigation budget we ingest the readily available 2022–2023 Excels
    and leave earlier years unavailable.
"""

from __future__ import annotations

import logging
import sys
from pathlib import Path

import requests
from openpyxl import load_workbook

from config import DATA_DIR, LOGS_DIR
from country_aliases import resolve_iso_from_name
from db import (
    get_cursor,
    load_geography_iso_map,
    normalize_country_name,
    upsert_indicator,
)

SOURCE = "heritage"
# Only years with a working direct Excel download (probed 2000–2023).
YEARS = [2022, 2023]
DOWNLOAD_URL_TMPL = (
    "https://indexdotnet.azurewebsites.net/index/excel/{year}/index{year}_data.xlsx"
)

# Column name variants across yearly workbooks → (indicator_code, label)
COMPONENT_FIELDS = [
    (("Business Freedom",), "heritage_business_freedom", "Business Freedom"),
    (("Trade Freedom",), "heritage_trade_freedom", "Trade Freedom"),
    (
        ("Investment Freedom ", "Investment Freedom"),
        "heritage_investment_freedom",
        "Investment Freedom",
    ),
    (("Financial Freedom",), "heritage_financial_freedom", "Financial Freedom"),
]


def configure_logging() -> None:
    LOGS_DIR.mkdir(parents=True, exist_ok=True)
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s %(levelname)s [%(name)s] %(message)s",
        handlers=[
            logging.StreamHandler(sys.stdout),
            logging.FileHandler(LOGS_DIR / "ingest_heritage.log", encoding="utf-8"),
        ],
    )


logger = logging.getLogger("ingest_heritage")


def cache_file(year: int) -> Path:
    return DATA_DIR / f"index{year}_data.xlsx"


def ensure_workbook(year: int) -> Path:
    path = cache_file(year)
    if path.exists() and path.stat().st_size > 0:
        logger.info("Using cached Heritage workbook: %s", path)
        return path
    url = DOWNLOAD_URL_TMPL.format(year=year)
    logger.info("Downloading Heritage data from %s", url)
    response = requests.get(url, timeout=120)
    response.raise_for_status()
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_bytes(response.content)
    logger.info("Cached %s bytes to %s", len(response.content), path)
    return path


def parse_numeric(raw) -> float | None:
    if raw is None or raw == "" or raw == "N/A" or raw == "#VALUE!":
        return None
    try:
        return float(raw)
    except (TypeError, ValueError):
        return None


def resolve_active_fields(header_index: dict[str, int], year: int) -> list[tuple[int, str, str]]:
    active: list[tuple[int, str, str]] = []
    score_col = f"{year} Score"
    if score_col in header_index:
        active.append((header_index[score_col], "heritage_overall", "Overall Score"))

    for aliases, code, label in COMPONENT_FIELDS:
        for col_name in aliases:
            if col_name in header_index:
                active.append((header_index[col_name], code, label))
                break
    return active


def ingest_year(year: int, iso_map: dict[str, str], cursor) -> tuple[int, list[str]]:
    path = ensure_workbook(year)
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        raise RuntimeError(f"Heritage workbook {year} is empty")

    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    header_index = {h: i for i, h in enumerate(headers) if h}
    logger.info("Heritage %s columns: %s", year, headers)

    name_col = header_index.get("Country Name")
    if name_col is None:
        raise RuntimeError(f"Country Name column not found in Heritage workbook {year}")

    active_fields = resolve_active_fields(header_index, year)
    if not active_fields:
        raise RuntimeError(f"No Heritage indicator columns found for {year}")

    url = DOWNLOAD_URL_TMPL.format(year=year)
    stored = 0
    unmatched: list[str] = []
    for row in rows[1:]:
        if not row or name_col >= len(row):
            continue
        country_name = row[name_col]
        if not country_name:
            continue
        iso = resolve_iso_from_name(str(country_name), normalize_country_name)
        if not iso or iso not in iso_map:
            unmatched.append(str(country_name))
            continue
        geo_id = iso_map[iso]
        for col_idx, code, label in active_fields:
            value = parse_numeric(row[col_idx] if col_idx < len(row) else None)
            upsert_indicator(
                cursor,
                geography_id=geo_id,
                source=SOURCE,
                indicator_code=code,
                indicator_name=label,
                value=value,
                unit="index_score",
                year=year,
                data_url=url,
            )
            stored += 1
    return stored, unmatched


def ingest() -> None:
    stored = 0
    unmatched_all: list[str] = []
    with get_cursor() as cursor:
        iso_map = load_geography_iso_map(cursor)
        for year in YEARS:
            y_stored, unmatched = ingest_year(year, iso_map, cursor)
            logger.info(
                "Year %s stored=%s unmatched_countries=%s",
                year,
                y_stored,
                len(unmatched),
            )
            stored += y_stored
            unmatched_all.extend(unmatched)

    logger.info(
        "Done stored=%s years=%s unmatched_total=%s "
        "(pre-2022 Excel downloads unavailable at known URL pattern)",
        stored,
        YEARS,
        len(unmatched_all),
    )
    if unmatched_all:
        unique = sorted(set(unmatched_all))
        logger.info("Unmatched Heritage countries: %s", unique)


if __name__ == "__main__":
    configure_logging()
    try:
        ingest()
        logger.info("ingest_heritage completed successfully")
    except Exception:
        logger.exception("ingest_heritage failed")
        sys.exit(1)
